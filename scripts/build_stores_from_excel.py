import json
import re
import sys
import time
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl

EXCEL_PATH = "SCRAPPING_DATABASE_SEPT 2025_AF_JMM+CARTIER+LINDBERG+SERENGETI+DITA+MAUIJIM+MYKITA+CHANEL+PORSCHE DESIGN+OAKLEY+JULBO.xlsx"
OUTPUT_PATH = "public/stores.json"
CACHE_PATH = "scripts/.geocode_cache.json"

BRAND_LABELS = {
    "PORSCHE DESIGN": "Porsche Design",
    "OAKLEY": "Oakley",
    "JULBO": "Julbo",
    "CHANEL": "Chanel",
    "MYKITA": "Mykita",
    "MAUI JIM": "Maui Jim",
    "DITA": "Dita",
    "SERENGETI": "Serengeti",
    "LINDBERG": "Lindberg",
    "JMM": "JMM",
    "CARTIER": "Cartier",
    "THOM BROWNE": "Thom Browne",
    "THIERRY LASRY": "Thierry Lasry",
    "MOSCOT": "Moscot",
}

LOWERCASE_PARTICLES = {"de", "du", "des", "la", "le", "les", "d", "l", "et", "sur", "en"}

BP_NUMBER_REGEX = re.compile(r"\bBP\s*\d+", re.IGNORECASE)
ZIP_REGEX = re.compile(r"\b(\d{5})\b")


def extract_expected_zip(address):
    # A "BP <digits>" postal-box reference earlier in the address is a
    # 5-digit run too and would otherwise be picked up as if it were the
    # real postal code — strip it first, then take the last remaining
    # 5-digit token (the real postal code consistently sits right before
    # the city name at the end of these addresses).
    cleaned = BP_NUMBER_REGEX.sub("", address)
    matches = ZIP_REGEX.findall(cleaned)
    return matches[-1] if matches else None


def expected_dept(zip_code):
    if not zip_code:
        return None
    return zip_code[:3] if zip_code.startswith(("97", "98")) else zip_code[:2]


def norm(value):
    return str(value).strip() if value is not None else ""


def titleize(text):
    if not text:
        return text
    if not text.isupper():
        return text
    words = text.lower().split(" ")
    out = []
    for i, word in enumerate(words):
        core = re.sub(r"[^a-zàâäéèêëïîôöùûüç']", "", word)
        if i > 0 and core in LOWERCASE_PARTICLES:
            out.append(word)
        else:
            out.append(word[:1].upper() + word[1:] if word else word)
    return " ".join(out)


def dedup_key(name, city):
    n = re.sub(r"[^A-Z0-9]", "", norm(name).upper())
    c = re.sub(r"[^A-Z0-9]", "", norm(city).upper())
    return (n, c)


def slugify(text):
    text = norm(text).lower()
    text = (
        text.replace("é", "e").replace("è", "e").replace("ê", "e").replace("ë", "e")
        .replace("à", "a").replace("â", "a").replace("ä", "a")
        .replace("ô", "o").replace("ö", "o")
        .replace("ù", "u").replace("û", "u").replace("ü", "u")
        .replace("ç", "c").replace("î", "i").replace("ï", "i")
    )
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "store"


# Geocoding used to be done by city name alone (`type=municipality`, no
# street or postal code). That collapsed every store in a city onto one
# identical point, and — worse — silently mis-resolved onto the wrong place
# whenever a town name is shared by more than one French commune (e.g.
# "Saint-Louis" in Haut-Rhin vs La Réunion, "Montreuil" in Seine-Saint-Denis
# vs Pas-de-Calais, "Les Angles" in Gard vs Pyrénées-Orientales — all
# confirmed wrong in the previously generated stores.json). The postal code
# is normally embedded in the free-text `address` field already, so
# geocoding the full address (falling back to the plain city name only if
# that yields nothing) resolves the correct commune.
def geocode_query(query_text, retries=3):
    """Raw geocode of one free-text query string against the BAN (French
    government address database, api-adresse.data.gouv.fr) — France-only by
    construction, so there is no risk of matching an address in another
    country. Returns (lat, lng, postcode) — the postcode lets callers
    cross-check the match against the postal code already stated in the
    source address (see resolve() below): a noisy query (store/mall name
    mixed into the address) can occasionally out-fuzzy-match onto an
    unrelated place elsewhere in France with a low relevance score, and the
    postcode is a cheap, reliable way to catch that."""
    query = urllib.parse.urlencode({"q": query_text, "limit": 1})
    url = f"https://api-adresse.data.gouv.fr/search/?{query}"
    last_error = None
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=15) as response:
                data = json.load(response)
            features = data.get("features") or []
            if not features:
                return None
            props = features[0]["properties"]
            lng, lat = features[0]["geometry"]["coordinates"]
            return (lat, lng, props.get("postcode"))
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            time.sleep(1.5 * (attempt + 1))
    print(f"  geocode failed for {query_text!r}: {last_error}", file=sys.stderr)
    return None


def build_full_query(address, city):
    addr = norm(address)
    city_n = norm(city)
    if not addr:
        return city_n
    if city_n and city_n.lower() not in addr.lower():
        return f"{addr}, {city_n}"
    return addr


def load_cache():
    try:
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            raw = json.load(f)
        return {k: (tuple(v) if v is not None else None) for k, v in raw.items()}
    except FileNotFoundError:
        return {}


def save_cache(cache):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f)


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

    stores = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [norm(h).lower() for h in row]
                continue
            if row[0] is None:
                continue
            d = dict(zip(header, row))
            country = norm(d.get("country"))
            if country.lower() != "france":
                continue
            name = norm(d.get("name"))
            city = norm(d.get("city") or d.get("city of industry"))
            if not name or not city:
                continue

            key = dedup_key(name, city)
            brand = BRAND_LABELS[sheet]
            if key not in stores:
                stores[key] = {
                    "name": titleize(name),
                    "city": titleize(city),
                    "address": norm(d.get("address")),
                    "phone": norm(d.get("phone") or d.get("phone number")),
                    "email": norm(d.get("email")),
                    "website": norm(d.get("website")),
                    "brands": [],
                }
            if brand not in stores[key]["brands"]:
                stores[key]["brands"].append(brand)

    print(f"Deduped France stores: {len(stores)}", file=sys.stderr)

    cache = load_cache()

    # One geocoding query per unique (address, city) — not per unique city —
    # so stores sharing a commune name but differing by street/postal code
    # each resolve independently instead of collapsing onto one point.
    # Each queued item also carries the plain city name (fallback query if
    # the full address doesn't match anything) and the postal code already
    # stated in the source address (used to cross-check the match).
    pending = {}
    for store in stores.values():
        query = build_full_query(store["address"], store["city"])
        key = query.lower()
        if key not in cache and key not in pending:
            expected_zip = extract_expected_zip(store["address"])
            pending[key] = (query, store["city"], expected_zip)

    print(f"Unique queries to geocode: {len(pending)} (cached: {len(stores) - len(pending)} stores)", file=sys.stderr)

    def resolve(query, city, expected_zip):
        expected = expected_dept(expected_zip)
        result = geocode_query(query)
        if result:
            _, _, postcode = result
            if expected and expected_dept(postcode) != expected:
                # Fuzzy-matched to a plausible-looking but wrong place (e.g.
                # a shop/mall name in the address pulled the match toward an
                # unrelated "Zone Commerciale" elsewhere in France) — the
                # postal code the source data already gives us disagrees
                # with where this landed, so don't trust it.
                result = None
        if not result and query.lower() != city.lower():
            result = geocode_query(city)
        if not result:
            return None
        lat, lng, _ = result
        return (lat, lng)

    done_count = 0
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {
            executor.submit(resolve, query, city, expected_zip): key
            for key, (query, city, expected_zip) in pending.items()
        }
        for future in as_completed(futures):
            key = futures[future]
            cache[key] = future.result()
            done_count += 1
            if done_count % 100 == 0:
                print(f"  geocoded {done_count}/{len(pending)}", file=sys.stderr)
                save_cache(cache)
    save_cache(cache)

    result = []
    used_ids = set()
    skipped = 0
    for store in stores.values():
        query_key = build_full_query(store["address"], store["city"]).lower()
        coords = cache.get(query_key)
        if not coords:
            skipped += 1
            continue
        lat, lng = coords

        base_id = slugify(f"{store['name']}-{store['city']}")
        store_id = base_id
        suffix = 2
        while store_id in used_ids:
            store_id = f"{base_id}-{suffix}"
            suffix += 1
        used_ids.add(store_id)

        entry = {
            "id": store_id,
            "name": store["name"],
            "address": store["address"] or f"{store['city']}, France",
            "city": store["city"],
            "country": "France",
            "lat": round(lat, 5),
            "lng": round(lng, 5),
            "brands": sorted(store["brands"]),
        }
        if store["phone"]:
            entry["phone"] = store["phone"]
        if store["email"]:
            entry["email"] = store["email"]
        if store["website"]:
            entry["website"] = store["website"]

        result.append(entry)

    result.sort(key=lambda s: (s["city"], s["name"]))

    print(f"Skipped (no geocode): {skipped}", file=sys.stderr)
    print(f"Final store count: {len(result)}", file=sys.stderr)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
        f.write("\n")


if __name__ == "__main__":
    main()
